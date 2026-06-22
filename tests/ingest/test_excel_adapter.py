"""Tests for Excel adapter."""

from __future__ import annotations

import tempfile

import openpyxl
import pytest

from spec_atlas.ingest.adapters.excel import ExcelAdapter
from spec_atlas.ingest.source_unit import SourceType


@pytest.fixture
def excel_file() -> str:
    """Create a test Excel file.

    Returns:
        Path to the created Excel file.
    """
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # Headers
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        ws["C1"] = "Status"

        # Data rows
        ws["A2"] = "Item1"
        ws["B2"] = 100
        ws["C2"] = "Active"

        ws["A3"] = "Item2"
        ws["B3"] = 200
        ws["C3"] = "Inactive"

        wb.save(f.name)
        wb.close()
        yield f.name


@pytest.fixture
def multi_sheet_excel() -> str:
    """Create a test Excel file with multiple sheets.

    Returns:
        Path to the created Excel file.
    """
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb = openpyxl.Workbook()

        # First sheet
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "Col1"
        ws1["B1"] = "Col2"
        ws1["A2"] = "Data1"
        ws1["B2"] = "Data2"

        # Second sheet
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "Name"
        ws2["B1"] = "Email"
        ws2["A2"] = "John"
        ws2["B2"] = "john@example.com"

        wb.save(f.name)
        wb.close()
        yield f.name


@pytest.mark.anyio
async def test_excel_adapter_parses_rows(excel_file: str) -> None:
    """Test that Excel adapter parses rows into SourceUnits."""
    adapter = ExcelAdapter("test.xlsx", excel_file)
    units = await adapter.ingest()

    assert len(units) == 2
    assert "Item1" in units[0].text
    assert "100" in units[0].text
    assert "Active" in units[0].text
    assert "Item2" in units[1].text
    assert "200" in units[1].text


@pytest.mark.anyio
async def test_excel_provenance(excel_file: str) -> None:
    """Test that Excel adapter creates correct provenance."""
    adapter = ExcelAdapter("test.xlsx", excel_file)
    units = await adapter.ingest()

    assert units[0].provenance.source_type == SourceType.EXCEL
    assert "TestSheet" in units[0].provenance.locator
    assert "row=2" in units[0].provenance.locator
    assert "sheet=" in units[0].provenance.locator

    assert "row=3" in units[1].provenance.locator


@pytest.mark.anyio
async def test_excel_structure(excel_file: str) -> None:
    """Test that Excel adapter captures row structure."""
    adapter = ExcelAdapter("test.xlsx", excel_file)
    units = await adapter.ingest()

    assert units[0].structure is not None
    assert "Name" in units[0].structure
    assert "Value" in units[0].structure
    assert units[0].structure["Name"] == "Item1"
    assert units[0].structure["Value"] == 100


@pytest.mark.anyio
async def test_excel_multi_sheet(multi_sheet_excel: str) -> None:
    """Test that Excel adapter handles multiple sheets."""
    adapter = ExcelAdapter("multi.xlsx", multi_sheet_excel)
    units = await adapter.ingest()

    assert len(units) == 2
    assert "Sheet1" in units[0].provenance.locator
    assert "Sheet2" in units[1].provenance.locator


@pytest.mark.anyio
async def test_excel_adapter_invalid_file() -> None:
    """Test that ExcelAdapter raises ValueError for invalid file."""
    adapter = ExcelAdapter("nonexistent.xlsx", "/nonexistent/path.xlsx")

    with pytest.raises(ValueError, match="Failed to parse Excel"):
        await adapter.ingest()


@pytest.mark.anyio
async def test_excel_adapter_source_id(excel_file: str) -> None:
    """Test that source_id is preserved in SourceUnits."""
    adapter = ExcelAdapter("my_source.xlsx", excel_file)
    units = await adapter.ingest()

    assert all(unit.source_id == "my_source.xlsx" for unit in units)
