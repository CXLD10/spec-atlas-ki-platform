"""Adapter for Excel spreadsheets (.xlsx)."""

from __future__ import annotations

import logging

import openpyxl

from spec_atlas.ingest.adapters.base import SourceAdapter
from spec_atlas.ingest.source_unit import Provenance, SourceType, SourceUnit

logger = logging.getLogger(__name__)


class ExcelAdapter(SourceAdapter):
    """Adapter for Excel spreadsheets."""

    def __init__(self, source_id: str, file_path: str):
        """Initialize Excel adapter.

        Args:
            source_id: Unique identifier (filename or source name).
            file_path: Absolute path to the Excel file.
        """
        super().__init__(source_id)
        self.file_path = file_path

    async def ingest(self) -> list[SourceUnit]:
        """Parse Excel workbook into row-level SourceUnits.

        Returns:
            List of SourceUnit, one per data row with provenance linking to sheet and row.

        Raises:
            ValueError: If Excel file cannot be opened or parsed.
        """
        units = []

        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Extract headers from first row
                headers = []
                for cell in ws[1]:
                    headers.append(cell.value if cell.value else "")

                # Skip empty header row
                if not any(headers):
                    continue

                # Process data rows (starting from row 2)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                    row_data = {}
                    row_values = []

                    for col_idx, cell in enumerate(row):
                        if col_idx < len(headers):
                            header = headers[col_idx]
                            value = cell.value
                            if header and value is not None:
                                row_data[header] = value
                                row_values.append(str(value))

                    # Skip empty rows
                    if not row_data:
                        continue

                    # Create SourceUnit for this row
                    text = " | ".join(row_values)

                    unit = SourceUnit(
                        source_id=self.source_id,
                        text=text,
                        structure=row_data,
                        provenance=Provenance(
                            source_type=SourceType.EXCEL,
                            locator=f"{self.source_id}:sheet={sheet_name}:row={row_idx}",
                            source_id=self.source_id,
                        ),
                    )
                    units.append(unit)

            wb.close()

        except Exception as e:
            raise ValueError(f"Failed to parse Excel {self.file_path}: {e}") from e

        if not units:
            logger.warning(f"Excel {self.source_id} contains no data rows")

        logger.info(f"Excel adapter ingested {len(units)} rows from {self.source_id}")
        return units
